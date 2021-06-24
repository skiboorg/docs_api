# coding utf-8

import ftputil
from django.template import loader, Context
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import generics
from .serializers import *
from .models import *
from .services import *
from datetime import datetime
import settings
from user.models import User
from django.core.mail import send_mail
from django.template.loader import render_to_string

class AddSubscribe(APIView):
    def post(self, request):
        MailSubscribe.objects.create(email=request.data.get('email'))
        return  Response(status=200)
class PayComplete(APIView):
    def post(self, request):
        data = request.data
        payment_id = data['object']['id']
        try:
            payment = PaymentObj.objects.get(ya_id=payment_id)
            if not payment.is_payed:
                payment.is_payed = True
                payment.order.is_payed = True
                payment.save()
                payment.order.save()
                print('sending to crm',payment.order)
                send_order_to_crm(payment.order)
            return Response(status=200)
        except:
             return Response(status=500)

class Test(APIView):

    def get(self, request):
        items = ItemImage.objects.all()
        for item in items:
            item.save()

        return Response(status=200)
#from openpyxl import load_workbook

# class Test(APIView):
#     def get(self, request):
#
#         cities = City.objects.all()
#         for c in cities:
#             City.objects.create(name=c.name,code=c.code,type_id=1)
#
#         return Response(status=200)

# class Test(APIView):
#     def get(self, request):
#         wb = load_workbook(filename='citie.xlsx')
#         sheet = wb.active
#
#         max_row = sheet.max_row
#
#         max_column = sheet.max_column
#         ii=0
#         for i in range(2, max_row + 1):
#             city_id = sheet.cell(row=i, column=1).value
#             city = sheet.cell(row=i, column=3).value
#             try:
#                 city_in_db = City.objects.get(name=city,type__is_office_cdek=True)
#                 city_in_db.code = city_id
#                 city_in_db.save()
#                 print('found')
#             except:
#
#                 print('not f')
#         print(ii)
#         return Response(status=200)

class GetUserOrders(generics.ListAPIView):
    serializer_class = OrderSerializer
    def get_queryset(self):
        return Order.objects.filter(client_id=self.request.query_params.get('user_id'))
class GetDelivery(generics.ListAPIView):
    serializer_class = DeliverySerializer
    queryset = DeliveryType.objects.all()


class GetItems(generics.ListAPIView):
    serializer_class = ItemTypeSiteMapSerializer
    queryset = ItemType.objects.all()

class GetBanners(generics.ListAPIView):
    serializer_class = BannerSerializer
    queryset = Banner.objects.all()


class GetCategories(generics.ListAPIView):
    serializer_class = CategorySerializer
    queryset = Category.objects.all()


class GetSubcategoryItems(generics.ListAPIView):
    serializer_class = SimpleItemSerializer
    def get_queryset(self):
        items = Item.objects.filter(subcategory__name_slug=self.request.query_params['subcategory_name_slug'], is_active=True)
        if items:
            return items
        else:
            return

class GetCollectionBySlug(generics.ListAPIView):
    serializer_class = CollectionWithItemsSerializer
    def get_queryset(self):
        return Collection.objects.filter(name_slug=self.request.query_params['name_slug'])


class GetHomeCollections(generics.ListAPIView):
    serializer_class = CollectionWithItemsSerializer
    def get_queryset(self):
        return Collection.objects.filter(is_show_at_home=True)

class GetNewItems(generics.ListAPIView):
    serializer_class = CollectionItemSerializer
    def get_queryset(self):
        return Item.objects.filter(is_new=True)


class GetItem(generics.RetrieveAPIView):
    serializer_class = ItemSerializer

    def get_object(self):
        base_item_slug = self.request.query_params['base_item_slug']
        item = Item.objects.get(name_slug=base_item_slug)
        return item

class GetRecommendedItems(generics.ListAPIView):
    serializer_class = RecItemSerializer

    def get_queryset(self):
        base_item_slug = self.request.query_params['base_item_slug']
        item = Item.objects.get(name_slug=base_item_slug)
        subcat = item.recommended_subcategory
        if subcat:
            items = subcat.subcategory_items.filter(is_active=True).exclude(id=item.id)[:3]
            return items
        else:
            return

class GetItem_old(APIView):
    def get(self,request):
        colors = []
        base_item_slug = self.request.query_params['base_item_slug']

        item = Item.objects.get(name_slug=base_item_slug)
        types = ItemType.objects.filter(item=item)

        for type in types:
            color_for_add = {
                "color_id": type.color.id,
                "color_hex": type.color.bg_color,
                "color_name": type.color.name
            }
            if not color_for_add in colors:
                colors.append(color_for_add)

        for color in colors:
            color["sizes"] = []
            color["images"] = []

        for type in types:
            for color in colors:
                if color["color_id"] == type.color.id:
                    size_to_add = {
                        "size_id": type.size.id,
                        "size_name": type.size.name,
                        "heights": []
                    }
                    if not size_to_add in color["sizes"]:
                        color["sizes"].append(size_to_add)
        for type in types:
            for color in colors:
                for size in color['sizes']:
                    if color['color_id'] == type.color.id and size['size_id'] == type.size.id:
                        height_to_add = {
                            "height_id": type.height.id,
                            "height_name": type.height.name,
                        }
                        if not height_to_add in size['heights']:
                            size['heights'].append(height_to_add)

        for color in colors:
            images = ItemImage.objects.filter(item=type.item, color_id=color['color_id'])
            for image in images:
                image_to_add = {
                    "image_id": image.id,
                    "image": image.image.url,
                    "image_thumb": image.image_thumb.url,
                }
                color['images'].append(image_to_add)
        result = {
            "item_data":ItemSerializer(item).data,
            "colors_data": colors
        }

        return Response(result,status=200)


class GetCart(APIView):
    def get(self,request):
        cart = check_if_cart_exists(request, self.request.query_params.get('session_id'))
        print(cart)
        serializer = CartSerializer(cart)
        return Response(serializer.data,status=200)

class CreateOrder(APIView):
    def post(self, request):
        data = request.data
        print(data)
        session_id = data.get('session_id')
        order_data = data.get('order')
        delivery_price = data.get('delivery_price')
        cart = check_if_cart_exists(request, session_id)
        pack_price = 0
        need_pack = False
        pack_type = order_data.get('pack_type')
        if pack_type == 'pack':
            pack_price = 300
            need_pack = True
        need_register = order_data.get('need_register')
        if need_register:
            user = User.objects.create_user(order_data.get('email'), '0000')
            user.fio = order_data.get('fio')
            user.phone = order_data.get('phone')
            user.save()

        new_order = Order.objects.create(
            payment=order_data.get('pay_type'),
            phone=order_data.get('phone'),
            email=order_data.get('email'),
            fio=order_data.get('fio'),
            street=order_data.get('street'),
            house=order_data.get('house'),
            flat=order_data.get('flat'),
            delivery_id=order_data.get('delivery_type'),
            city_id=order_data.get('delivery_city') if order_data.get('delivery_city') else None,
            comment=order_data.get('comment'),
            promo_code=cart.promo_code,
            delivery_price=delivery_price,
            order_code=''.join(choices(string.ascii_lowercase + string.digits, k=8)),
            weight=cart.weight,
            is_need_pack=need_pack,
            total_price=cart.total_price + pack_price
        )

        if cart.client:
            new_order.client = cart.client
        else:
            new_order.guest = cart.guest

        new_order.save()

        # msg_html = render_to_string('new_order.html', {'order': new_order,
        #                                                'items': cart.items.all()})
        # send_mail('Ваш заказ', None, 'noreply@docsuniform.ru', [new_order.email,'info@docsuniform.ru'],
        #           fail_silently=False, html_message=msg_html)

        for item in cart.items.all():
            new_order_item = OrderItem.objects.create(item_type=item.item_type,quantity=item.quantity)
            new_order.items.add(new_order_item)
            item.delete()

        cart.promo_code = None
        cart.save()
        pay_url = None

        if order_data.get('pay_type')=='online':
            pay_url = pay_request(new_order)

        else:
            send_order_to_crm(new_order)
        return Response({'order_code': new_order.order_code, 'pay_url': pay_url}, status=200)


class ApplyPromo(APIView):
    def post(self, request):
        data = request.data
        try:
            promo = PromoCode.objects.get(code=data.get('code'))
            cart = check_if_cart_exists(request, data.get('session_id'))
            cart.promo_code = promo
            cart.save()
            return Response({'status': True}, status=200)
        except:

            return Response({'status':False}, status=200)

class MinusQuantity(APIView):
    def post(self, request):
        data = request.data
        item = CartItem.objects.get(id=data.get('item_id'))
        if item.quantity > 1:
            item.quantity -= 1
            item.save()
        else:
            item.delete()
        return Response(status=200)

class PlusQuantity(APIView):
    def post(self, request):
        data = request.data
        item = CartItem.objects.get(id=data.get('item_id'))
        item.quantity += 1
        item.save()
        return Response(status=200)

class DeleteItem(APIView):
    def post(self, request):
        data = request.data
        item = CartItem.objects.get(id=data.get('item_id'))
        item.delete()
        return Response(status=200)


class AddToCart(APIView):
    def post(self, request):
        data = request.data
        print(request.data)
        user = None
        guest = None
        cart = None
        adding_item_type = ItemType.objects.get(id=data.get('item_id'))
        cart = check_if_cart_exists(request, data.get('session_id'))
        add_item_to_cart(cart, adding_item_type)
        return Response(status=200)


def catalog_feed(request,):
    template_vars = {}
    template_vars['categories'] = Category.objects.all()
    template_vars['item_types'] = ItemType.objects.all()
    template_vars['base_url'] = settings.BASE_URL
    template_vars['created'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    t = loader.get_template('catalog.xml')
    c = template_vars
    return HttpResponse(t.render(c),content_type='text/xml')


class CheckOstatok(APIView):
    def get(self, request):
        from lxml import etree
        with ftputil.FTPHost('185.92.148.221', settings.FTP_USER, settings.FTP_PASSWORD) as host:
            names = host.listdir(host.curdir)
            for name in names:
                if host.path.isfile(name):
                    host.download(name, name)
        tree = etree.parse('vigruzkaostatok.xml')
        root = tree.getroot()
        print(root)
        items = 0
        for element in root:
            print(element)
            item_id = element.find("item_id").text
            quantity = element.find("goods").text
            print(item_id)
            try:
                item = ItemType.objects.get(id_1c=item_id)
                item.quantity = int(quantity)
                item.save()
                items += 1
            except:
                pass
        return Response({'Обновлено остатков':items},status=200)

class CheckFtp(APIView):

    def get(self, request):
        from lxml import etree
        with ftputil.FTPHost('185.92.148.221', settings.FTP_USER, settings.FTP_PASSWORD) as host:
            names = host.listdir(host.curdir)
            for name in names:
                if host.path.isfile(name):
                    host.download(name, name)
        tree = etree.parse('tovar.xml')
        root = tree.getroot()
        new_items = 0
        updated_items = 0
        new_items_types = 0
        updated_items_types = 0
        for element in root:
            item_id = element.find("basic_item").text
            item_name = element.find("name").text
            item_price = element.find("price").text
            item, created = Item.objects.get_or_create(id_1c=item_id)
            print(item_name)
            if created:
                new_items += 1
            else:
                updated_items += 1
            item.name = item_name
            item.price = int(item_price)
            item.save()

        tree = etree.parse('vigruzka.xml')
        root = tree.getroot()

        for element in root:
            basic_item_id = element.find("basic_item").text
            base_item = None
            try:
                base_item = Item.objects.get(id_1c=basic_item_id)
            except Item.DoesNotExist:
                print('BASE ITEM NOT EXIST. ABORT')
                base_item = None


            color_id = element.find("color").text
            color, created = ItemColor.objects.get_or_create(id_1c=color_id)
            if created:
                color.add_id = color_id
                color.name = 'Новый цвет'
                color.save()

            size_name = element.find("size").text
            size, created = ItemSize.objects.get_or_create(name=size_name)
            if created:
                size.name = size_name
                size.save()

            height_name = element.find("height").text
            height, created = ItemHeight.objects.get_or_create(name=height_name)
            if created:
                height.name = height_name
                height.save()

            add_id = element.find("add").text
            if not add_id:
                add_id = 0

            mod, created = ItemModification.objects.get_or_create(id_1c=add_id)
            if created:
                mod.id_1c = add_id
                mod.name = 'Новая модификация'
                mod.save()

            cloth_id = element.find("cloth").text
            material, created = ItemMaterial.objects.get_or_create(id_1c=cloth_id)
            if created:
                material.id_1c = cloth_id
                material.name = 'Новая ткань'
                material.save()

            item_id = element.find("item_id").text

            try:
                bar_code = element.find("barcode").text
            except:
                bar_code = ''
            if base_item:
                item_type, created = ItemType.objects.get_or_create(id_1c=item_id)
                if created:
                    new_items_types += 1
                item_type.item = base_item
                item_type.color = color
                item_type.bar_cade = bar_code
                item_type.size = size
                item_type.height = height
                item_type.material = material
                item_type.modification = mod
                item_type.save()
                updated_items_types += 1

        return Response({'Создано базовых товаров':new_items,
                         'Обновлено базовый товаров':updated_items,
                         'Создано типов товаров':new_items_types,
                         'Обновлено типов товаров':updated_items_types,
                         },
                        status=200)



class CalculateDelivery(APIView):
    def get(self, request):
        print(self.request.query_params)
        cdek_type = self.request.query_params.get('cdek_type')
        city_code = self.request.query_params.get('city_code')
        weight = self.request.query_params.get('weight')
        access_token = checkCdekToken()
        print(city_code)
        price = calculateDelivery(access_token, city_code, weight, cdek_type)
        return Response({'delivery_price':price}, status=200)


class OrderPayed(APIView):
    def post(self,request):
        data = request.data
        payment_id = data.get('pay_id')
        print(payment_id)
        try:
            payment = PaymentObj.objects.get(pay_id=payment_id)
            if not payment.is_payed:
                payment.is_payed = True
                payment.order.is_payed = True
                payment.save()
                print('sending to crm',payment.order)
                send_order_to_crm(payment.order)
            return Response(status=200)
        except:
             return Response(status=500)

class FillC(APIView):

    def get(self,request):
        from openpyxl import load_workbook
        wb = load_workbook(filename='off.xlsx')
        sheet = wb.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        delivery = DeliveryType.objects.get(is_office_cdek=True)
        for i in range(1, max_row + 1):
            city = sheet.cell(row=i, column=2).value
            print(city)
            c, cret = City.objects.get_or_create(type=delivery,name=city,code=0)
            if cret:
                print('double')
        for i in range(1, max_row + 1):
            city_name = sheet.cell(row=i, column=2).value
            city = City.objects.get(name=city_name)
            code = sheet.cell(row=i, column=1).value
            addr = sheet.cell(row=i, column=4).value
            CdekOffice.objects.create(city=city,office_id=code,address=addr)
            print(addr)



        return Response(status=200)
